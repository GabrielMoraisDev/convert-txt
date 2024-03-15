import openpyxl
from flask import Flask, render_template, request, send_file
import io
import time

app = Flask(__name__)

def process_text_file(file):
<<<<<<< HEAD
    decoded_lines = []
    line_count = 0
    partial_data = []

    for line in file:
        decoded_line = line.decode('utf-8').strip().split("  ")
        decoded_lines.append(decoded_line)
        line_count += 1
=======
    try:
        lines = file.read().decode('utf-8').splitlines()
        print("Decodificação utf-8 bem sucedida.")
    except UnicodeDecodeError:
        # Se a decodificação utf-8 falhar, tenta com outra codificação
        lines = file.read().decode('latin-1').splitlines()
        print("Decodificação utf-8 falhou. Tentando com latin-1.")
        
    data = [line.strip().split("  ") for line in lines]
    return data
>>>>>>> parent of ca737f1 (Update)

        if line_count == 100:
            partial_data.append(decoded_lines)
            decoded_lines = []
            line_count = 0
            time.sleep(1)  # Intervalo de 1 segundo entre cada 100 linhas

    # Adiciona o restante dos dados, caso existam
    if decoded_lines:
        partial_data.append(decoded_lines)

    return partial_data

def create_excel_file(data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active

    for rows in data:
        for row in rows:
            ws.append(row)

    wb.save(filename)

@app.route('/')
def index():
    return render_template('upload.html')

@app.route('/upload', methods=['POST'])
def upload():
    if request.method == 'POST':
        file = request.files['file']
        if file:
<<<<<<< HEAD
            partial_data = process_text_file(file)
            if partial_data:
                output_files = []
                for i, data in enumerate(partial_data):
                    filename = f'converted_{i+1}.xlsx'
                    create_excel_file(data, filename)
                    output_files.append(filename)

                # Concatenando os arquivos Excel
                combined_wb = openpyxl.Workbook()
                combined_ws = combined_wb.active
                for filename in output_files:
                    wb = openpyxl.load_workbook(filename)
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        combined_ws.append(row)
                    wb.close()
                combined_filename = 'combined.xlsx'
                combined_wb.save(combined_filename)

                # Fazendo download do arquivo combinado
=======
            data = process_text_file(file)
            if data:
                wb = create_excel_file(data)
                
                # Salva o arquivo Excel em memória
>>>>>>> parent of ca737f1 (Update)
                output = io.BytesIO()
                with open(combined_filename, 'rb') as f:
                    output.write(f.read())
                output.seek(0)
<<<<<<< HEAD

                return send_file(output, as_attachment=True, download_name='combined.xlsx')
=======
                
                # Retorna o arquivo Excel como um anexo para download
                return send_file(output, as_attachment=True, download_name='converted.xlsx')
>>>>>>> parent of ca737f1 (Update)
            else:
                return 'Os dados do arquivo estão vazios.'
    return 'Error processing file.'

if __name__ == '__main__':
    app.run(debug=True)
